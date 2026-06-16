import pathlib

from openpyxl import load_workbook

from app.runner.executor import run_workflow

SAMPLES = pathlib.Path(__file__).resolve().parent.parent / "samples"


def test_load_csv_then_export_xlsx():
    sales = SAMPLES / "sales.csv"
    workflow = [
        {"block_id": "load_csv", "params": {"path": str(sales)}, "input_bindings": {}},
        {
            "block_id": "export_xlsx",
            "params": {"filename": "sales.xlsx", "sheet_name": "Raw"},
            "input_bindings": {"data": "0.data"},
        },
    ]

    result = run_workflow(workflow)

    assert result["status"] == "succeeded", result.get("logs")
    final = result["step_outputs"][-1]
    out_path = pathlib.Path(final["file"])
    assert out_path.exists(), f"expected output xlsx at {out_path}"
    assert out_path.stat().st_size > 0

    wb = load_workbook(out_path)
    assert "Raw" in wb.sheetnames
    ws = wb["Raw"]
    # header + 12 data rows
    assert ws.max_row == 13
    assert ws.cell(row=1, column=1).value == "region"


def test_missing_block_returns_failed():
    workflow = [
        {"block_id": "definitely_not_a_block", "params": {}, "input_bindings": {}},
    ]
    result = run_workflow(workflow)
    assert result["status"] == "failed"
    assert result["failed_at"] == 0


def test_candidate_block_materializes_and_runs():
    sales = SAMPLES / "sales.csv"
    candidate_code = """
import argparse, json, pathlib
import pandas as pd

ap = argparse.ArgumentParser()
ap.add_argument('--workdir', required=True)
ap.add_argument('--params', required=True)
ap.add_argument('--inputs', required=True)
ap.add_argument('--outputs', required=True)
args = ap.parse_args()

workdir = pathlib.Path(args.workdir)
params = json.loads(pathlib.Path(args.params).read_text())
inputs = json.loads(pathlib.Path(args.inputs).read_text())

df = pd.read_parquet(inputs['data'])
col = params['column']
df['size_bucket'] = df[col].apply(lambda q: 'small' if q < 6 else 'medium' if q < 10 else 'large')

out = workdir / 'data.parquet'
df.to_parquet(out, index=False)
pathlib.Path(args.outputs).write_text(json.dumps({'data': str(out)}))
print('size_bucket: wrote', len(df), 'rows')
"""
    workflow = [
        {"block_id": "load_csv", "params": {"path": str(sales)}, "input_bindings": {}},
        {
            "block_id": "size_bucket",
            "params": {"column": "quantity"},
            "input_bindings": {"data": "0.data"},
        },
        {
            "block_id": "export_xlsx",
            "params": {"filename": "bucketed.xlsx"},
            "input_bindings": {"data": "1.data"},
        },
    ]
    candidate_blocks = {
        "size_bucket": {
            "code": candidate_code,
            "name": "Add size bucket",
            "description": "Categorize quantity into small/medium/large.",
            "params_schema": {"column": {"type": "string", "required": True}},
            "input_schema": {"data": {"type": "dataframe", "required": True}},
            "output_schema": {"data": {"type": "dataframe"}},
        }
    }
    result = run_workflow(workflow, candidate_blocks=candidate_blocks)
    assert result["status"] == "succeeded", result.get("logs")
    out = pathlib.Path(result["step_outputs"][-1]["file"])
    assert out.exists()

    wb = load_workbook(out)
    ws = wb["Sheet1"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "size_bucket" in header


def test_bad_csv_path_propagates_failure():
    workflow = [
        {"block_id": "load_csv", "params": {"path": "/tmp/no-such-file.csv"}, "input_bindings": {}},
    ]
    result = run_workflow(workflow)
    assert result["status"] == "failed"
    assert result["failed_at"] == 0
